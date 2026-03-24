"""
Adhik Maas Daura Seva – store submissions in PostgreSQL.

Endpoints
─────────
User-facing
  GET  /adhik-maas/areas                       list of Pune areas
  GET  /adhik-maas/my-submission?user_id=X     current user's submission
  POST /adhik-maas/submit                      new submission (one per user)
  PUT  /adhik-maas/my-submission               user edits own submission (toggle-gated)

Admin
  GET  /adhik-maas/submissions                 all submissions with full user info
  PUT  /adhik-maas/submissions/<id>            update any submission
  GET  /adhik-maas/summary                     aggregated stats + permutation combinations
  PUT  /adhik-maas/submissions/<id>/shortlist  shortlist or un-shortlist a user
  GET  /adhik-maas/shortlisted                 list shortlisted users
  PUT  /adhik-maas/submissions/<id>/finalize   finalize or un-finalize a user
  GET  /adhik-maas/finalized                   list finalized users
  GET  /adhik-maas/map-data                    map: id, name, seva_type, lat, lon
"""

import os
import logging
from collections import defaultdict
from datetime import datetime
from flask import Blueprint, request, jsonify

router = Blueprint("adhik_maas", __name__)

SUPER_ADMIN_MOBILE = os.getenv("SUPER_ADMIN_MOBILE", "1234567890")

# Kept as a last-resort fallback only; the real source of truth is the DB.
_FALLBACK_AREAS: list[str] = []


def _get_area_lookup() -> dict[str, dict]:
    """
    Return {area_name_lower: {area_name, route_number, route_name, pin_code}}
    from the DB.  Used to validate areas and enrich submissions.
    """
    from model import AdhikMaasArea
    rows = AdhikMaasArea.query.filter_by(is_active=True).all()
    return {
        r.area_name.lower(): {
            "area_name":    r.area_name,
            "route_number": r.route_number,
            "route_name":   r.route_name,
            "pin_code":     r.pin_code,
        }
        for r in rows
    }


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _require_admin():
    """Return (error_response, status_code) if not admin, else (None, None)."""
    from model import User
    data = request.get_json(silent=True) or {}
    admin_user_id = request.args.get("admin_user_id") or data.get("admin_user_id")
    admin_mobile  = request.args.get("admin_mobile")  or data.get("admin_mobile")

    if admin_mobile and str(admin_mobile).strip() == SUPER_ADMIN_MOBILE:
        return None, None
    if not admin_user_id:
        return jsonify({"error": "admin_user_id or admin_mobile required"}), 401
    try:
        admin_user_id = int(admin_user_id)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid admin_user_id"}), 401
    user = User.query.get(admin_user_id)
    if not user or not getattr(user, "isadmin", False):
        return jsonify({"error": "Admin access required"}), 403
    return None, None


def _parse_seva_flags(seva_preference: str) -> dict:
    """
    Derive boolean flags + time slot from the pipe-separated seva_preference
    string the frontend sends (e.g. "padyapuja|seva|afternoon|shejarti").
    Also handles legacy values from the old three-option model.
    """
    parts = set((seva_preference or "").split("|"))
    has_padyapuja       = "padyapuja"  in parts or seva_preference == "padya_puja"
    has_seva_mahaprasad = "seva"       in parts or seva_preference == "abhishek_mahaprasad"
    has_shejarti        = "shejarti"   in parts or seva_preference == "shejarti_kakad_aarti"
    seva_time = next(
        (p for p in ["afternoon", "evening", "any"] if p in parts),
        "any" if seva_preference == "abhishek_mahaprasad" else None,
    )
    return {
        "has_padyapuja":       has_padyapuja,
        "has_seva_mahaprasad": has_seva_mahaprasad,
        "seva_time":           seva_time,
        "has_shejarti":        has_shejarti,
    }


def _apply_flags(submission, seva_preference: str):
    """Write parsed flags onto a submission ORM object in-place."""
    flags = _parse_seva_flags(seva_preference)
    submission.has_padyapuja       = flags["has_padyapuja"]
    submission.has_seva_mahaprasad = flags["has_seva_mahaprasad"]
    submission.seva_time           = flags["seva_time"]
    submission.has_shejarti        = flags["has_shejarti"]


def _submission_dict(s, u=None):
    """Serialise an AdhikMaasSubmission + optional User into a dict."""
    return {
        "id":                   s.id,
        "user_id":              s.user_id,
        "user_name":            f"{u.first_name} {u.last_name}".strip() if u else "",
        "user_mobile":          getattr(u, "mobile_number", None) if u else None,
        "user_zone_code":       getattr(u, "zone_code", None)     if u else None,
        "user_area":            getattr(u, "area", None)          if u else None,
        "flat_no":              getattr(u, "flat_no", None)       if u else None,
        "full_address":         getattr(u, "full_address", None)  if u else None,
        "landmark":             getattr(u, "landmark", None)      if u else None,
        "city":                 getattr(u, "city", None)          if u else None,
        "state":                getattr(u, "state", None)         if u else None,
        "pincode":              getattr(u, "pincode", None)       if u else None,
        # raw
        "seva_preference":      s.seva_preference,
        "seva_label":           getattr(s, "seva_label", None),
        "area":                 s.area,
        "route_number":         getattr(s, "route_number", None),
        "route_name":           getattr(s, "route_name", None),
        "pin_code":             getattr(s, "pin_code", None),
        # structured flags
        "has_padyapuja":        bool(getattr(s, "has_padyapuja", False)),
        "has_seva_mahaprasad":  bool(getattr(s, "has_seva_mahaprasad", False)),
        "seva_time":            getattr(s, "seva_time", None),
        "has_shejarti":         bool(getattr(s, "has_shejarti", False)),
        # admin-assigned schedule & confirmed seva
        "route_date":           getattr(s, "route_date").isoformat() if getattr(s, "route_date", None) else None,
        "final_seva":           getattr(s, "final_seva", None),
        # workflow
        "is_shortlisted":       bool(getattr(s, "is_shortlisted", False)),
        "shortlisted_at":       getattr(s, "shortlisted_at").isoformat() if getattr(s, "shortlisted_at", None) else None,
        "is_finalized":         bool(getattr(s, "is_finalized", False)),
        "finalized_at":         getattr(s, "finalized_at").isoformat()   if getattr(s, "finalized_at", None)   else None,
        "admin_notes":          getattr(s, "admin_notes", None),
        "submitted_at":         s.submitted_at.isoformat()    if s.submitted_at    else None,
    }


# ─── Public endpoints ─────────────────────────────────────────────────────────

@router.route("/adhik-maas/areas", methods=["GET"])
def get_areas():
    """
    Return active areas grouped by route.
    Optional query param: ?pin_code=411014  → filters to areas matching that pin.
    Optional query param: ?q=baner          → filters by area name substring.

    Response:
    {
      "routes": [
        {
          "route_number": "1A",
          "route_name": "North-East ...",
          "areas": [
            { "area_name": "Vishrantwadi", "pin_code": "411015" },
            ...
          ]
        }
      ],
      "flat": ["Vishrantwadi", ...]   ← all area names, for simple text search
    }
    """
    from model import AdhikMaasArea
    from collections import OrderedDict

    pin_filter = (request.args.get("pin_code") or "").strip()
    q_filter   = (request.args.get("q") or "").strip().lower()

    try:
        query = AdhikMaasArea.query.filter_by(is_active=True)
        if pin_filter:
            query = query.filter(AdhikMaasArea.pin_code == pin_filter)
        rows = query.order_by(AdhikMaasArea.route_number, AdhikMaasArea.sort_order).all()

        if q_filter:
            rows = [r for r in rows if q_filter in r.area_name.lower() or q_filter in r.pin_code]

        route_map: OrderedDict = OrderedDict()
        for r in rows:
            key = r.route_number
            if key not in route_map:
                route_map[key] = {"route_number": r.route_number, "route_name": r.route_name, "areas": []}
            route_map[key]["areas"].append({"area_name": r.area_name, "pin_code": r.pin_code})

        return jsonify({
            "routes": list(route_map.values()),
            "flat":   [r.area_name for r in rows],
        }), 200
    except Exception as e:
        logging.exception("get_areas error: %s", e)
        return jsonify({"error": "Failed to fetch areas"}), 500


@router.route("/adhik-maas/my-submission", methods=["GET"])
def get_my_submission():
    user_id = request.args.get("user_id")
    if not user_id:
        return jsonify({"error": "user_id is required"}), 400
    try:
        user_id = int(user_id)
    except (ValueError, TypeError):
        return jsonify({"error": "Invalid user_id"}), 400

    try:
        from model import AdhikMaasSubmission
        s = AdhikMaasSubmission.query.filter_by(user_id=user_id).first()
        if not s:
            return jsonify({"submitted": False}), 200

        return jsonify({
            "submitted":           True,
            "id":                  s.id,
            "seva_preference":     s.seva_preference,
            "seva_label":          s.seva_label,
            "area":                s.area,
            "route_number":        getattr(s, "route_number", None),
            "route_name":          getattr(s, "route_name", None),
            "pin_code":            getattr(s, "pin_code", None),
            "has_padyapuja":       bool(getattr(s, "has_padyapuja", False)),
            "has_seva_mahaprasad": bool(getattr(s, "has_seva_mahaprasad", False)),
            "seva_time":           getattr(s, "seva_time", None),
            "has_shejarti":        bool(getattr(s, "has_shejarti", False)),
            "submitted_at":        s.submitted_at.isoformat() if s.submitted_at else None,
        }), 200
    except Exception as e:
        logging.exception("get_my_submission error: %s", e)
        return jsonify({"error": "Failed to fetch submission"}), 500


@router.route("/adhik-maas/submit", methods=["POST"])
def submit_adhik_maas():
    from model import db, AdhikMaasSubmission
    try:
        data = request.get_json(force=True, silent=True) or {}
    except Exception:
        data = {}

    user_id         = data.get("user_id")
    seva_preference = data.get("seva_preference")
    area            = (data.get("area") or "").strip()

    if not user_id:
        return jsonify({"error": "user_id is required"}), 400
    if not seva_preference:
        return jsonify({"error": "seva_preference is required"}), 400
    if not area:
        return jsonify({"error": "area is required"}), 400

    area_lookup = _get_area_lookup()
    area_info   = area_lookup.get(area.lower())
    if not area_info:
        return jsonify({"error": "area must be one of the allowed areas"}), 400

    try:
        existing = AdhikMaasSubmission.query.filter_by(user_id=int(user_id)).first()
        if existing:
            return jsonify({
                "error": "You have already submitted. Only one submission per user is allowed.",
                "already_submitted": True,
            }), 409
    except Exception:
        pass

    submitted_at = data.get("submitted_at")
    try:
        submitted_at = datetime.fromisoformat(submitted_at.replace("Z", "+00:00")) if submitted_at else datetime.utcnow()
    except (ValueError, TypeError, AttributeError):
        submitted_at = datetime.utcnow()

    try:
        entry = AdhikMaasSubmission(
            user_id         = int(user_id),
            seva_preference = seva_preference,
            seva_label      = data.get("seva_label", "") or "",
            area            = area_info["area_name"],
            route_number    = area_info["route_number"],
            route_name      = area_info["route_name"],
            pin_code        = area_info["pin_code"],
            submitted_at    = submitted_at,
        )
        _apply_flags(entry, seva_preference)
        db.session.add(entry)
        db.session.commit()
        return jsonify({"message": "Submission saved", "id": entry.id}), 201
    except Exception as e:
        db.session.rollback()
        logging.exception("adhik_maas submit error: %s", e)
        return jsonify({"error": "Failed to save submission"}), 500


@router.route("/adhik-maas/my-submission", methods=["PUT"])
def update_my_submission():
    from model import db, AdhikMaasSubmission, FeatureToggle

    toggle = FeatureToggle.query.filter_by(toggle_name="allow_adhik_maas_edit").first()
    if not toggle or not toggle.toggle_enabled:
        return jsonify({"error": "Editing submissions is currently disabled by admin."}), 403

    try:
        data = request.get_json(force=True, silent=True) or {}
    except Exception:
        data = {}

    user_id         = data.get("user_id")
    seva_preference = data.get("seva_preference")
    area            = (data.get("area") or "").strip()

    if not user_id:
        return jsonify({"error": "user_id is required"}), 400
    if not seva_preference:
        return jsonify({"error": "seva_preference is required"}), 400
    if not area:
        return jsonify({"error": "area is required"}), 400

    area_lookup = _get_area_lookup()
    area_info   = area_lookup.get(area.lower())
    if not area_info:
        return jsonify({"error": "area must be one of the allowed areas"}), 400

    try:
        submission = AdhikMaasSubmission.query.filter_by(user_id=int(user_id)).first()
        if not submission:
            return jsonify({"error": "No existing submission found for this user."}), 404

        submission.seva_preference = seva_preference
        submission.seva_label      = data.get("seva_label", "") or ""
        submission.area            = area_info["area_name"]
        submission.route_number    = area_info["route_number"]
        submission.route_name      = area_info["route_name"]
        submission.pin_code        = area_info["pin_code"]
        _apply_flags(submission, seva_preference)
        db.session.commit()
        return jsonify({"message": "Submission updated", **_submission_dict(submission)}), 200
    except Exception as e:
        db.session.rollback()
        logging.exception("adhik_maas user update error: %s", e)
        return jsonify({"error": "Failed to update submission"}), 500


# ─── Admin: list / update ──────────────────────────────────────────────────────

@router.route("/adhik-maas/submissions", methods=["GET"])
def list_submissions_admin():
    """[Admin] All submissions with full user info and workflow status."""
    err, status = _require_admin()
    if err is not None:
        return err, status
    try:
        from model import AdhikMaasSubmission, User
        submissions = AdhikMaasSubmission.query.order_by(AdhikMaasSubmission.submitted_at.desc()).all()
        out = [_submission_dict(s, User.query.get(s.user_id)) for s in submissions]
        return jsonify({"submissions": out, "total": len(out)}), 200
    except Exception as e:
        logging.exception("list_submissions_admin error: %s", e)
        return jsonify({"error": "Failed to fetch submissions"}), 500


@router.route("/adhik-maas/submissions/<int:submission_id>", methods=["PUT"])
def update_submission_admin(submission_id):
    """[Admin] Update seva preference / area / notes for any submission."""
    err, status = _require_admin()
    if err is not None:
        return err, status
    from model import db, AdhikMaasSubmission

    data            = request.get_json(force=True, silent=True) or {}
    seva_preference = data.get("seva_preference")
    seva_label      = data.get("seva_label", "")
    area            = (data.get("area") or "").strip()
    admin_notes     = data.get("admin_notes")
    route_date_str  = data.get("route_date")      # "YYYY-MM-DD" or None
    final_seva      = data.get("final_seva")       # confirmed seva label

    submission = AdhikMaasSubmission.query.get(submission_id)
    if not submission:
        return jsonify({"error": "Submission not found"}), 404

    if seva_preference is not None:
        submission.seva_preference = seva_preference
        _apply_flags(submission, seva_preference)
    if seva_label is not None:
        submission.seva_label = seva_label
    if area:
        area_lookup = _get_area_lookup()
        area_info   = area_lookup.get(area.lower())
        if not area_info:
            return jsonify({"error": "area must be one of the allowed areas"}), 400
        submission.area         = area_info["area_name"]
        submission.route_number = area_info["route_number"]
        submission.route_name   = area_info["route_name"]
        submission.pin_code     = area_info["pin_code"]
    if admin_notes is not None:
        submission.admin_notes = admin_notes
    if route_date_str is not None:
        from datetime import date as date_type
        try:
            submission.route_date = date_type.fromisoformat(route_date_str) if route_date_str else None
        except ValueError:
            return jsonify({"error": "route_date must be YYYY-MM-DD"}), 400
    if final_seva is not None:
        submission.final_seva = final_seva if final_seva else None

    try:
        db.session.commit()
        return jsonify({"message": "Submission updated", **_submission_dict(submission)}), 200
    except Exception as e:
        db.session.rollback()
        logging.exception("adhik_maas admin update error: %s", e)
        return jsonify({"error": "Failed to update submission"}), 500


# ─── Admin: summary / permutation combinations ────────────────────────────────

@router.route("/adhik-maas/summary", methods=["GET"])
def get_summary_admin():
    """
    [Admin] Aggregated statistics and every permutation combination that
    actually exists in the data, so the admin can see which preference
    bundles are popular and plan assignments accordingly.

    Response shape:
    {
      "total": N,
      "shortlisted": N,
      "finalized": N,
      "seva_counts": {
        "padyapuja": N,
        "seva_mahaprasad": N,
        "shejarti": N
      },
      "time_counts": { "afternoon": N, "evening": N, "any": N, "none": N },
      "area_counts": { "<area>": N, ... },
      "combinations": [
        {
          "has_padyapuja": true,
          "has_seva_mahaprasad": true,
          "seva_time": "afternoon",
          "has_shejarti": false,
          "count": N,
          "users": [ { "id", "user_id", "user_name", "area", ... } ]
        },
        ...
      ]
    }
    """
    err, status = _require_admin()
    if err is not None:
        return err, status

    try:
        from model import AdhikMaasSubmission, User

        all_subs = AdhikMaasSubmission.query.order_by(AdhikMaasSubmission.area).all()

        # ── totals
        total       = len(all_subs)
        shortlisted = sum(1 for s in all_subs if s.is_shortlisted)
        finalized   = sum(1 for s in all_subs if s.is_finalized)

        # Re-derive flags from seva_preference for every submission so that old
        # records with NULL boolean columns are counted correctly.
        def _flags(s):
            return _parse_seva_flags(s.seva_preference)

        # ── seva breakdown
        seva_counts = {
            "padyapuja":       sum(1 for s in all_subs if _flags(s)["has_padyapuja"]),
            "seva_mahaprasad": sum(1 for s in all_subs if _flags(s)["has_seva_mahaprasad"]),
            "shejarti":        sum(1 for s in all_subs if _flags(s)["has_shejarti"]),
        }

        # ── time breakdown
        time_counts = defaultdict(int)
        for s in all_subs:
            key = _flags(s)["seva_time"] or "none"
            time_counts[key] += 1

        # ── area breakdown
        area_counts = defaultdict(int)
        for s in all_subs:
            area_counts[s.area or "Unknown"] += 1

        # ── route breakdown
        route_counts: dict = {}
        for s in all_subs:
            key = s.route_number or "Unknown"
            if key not in route_counts:
                route_counts[key] = {"route_number": key, "route_name": s.route_name or "", "count": 0}
            route_counts[key]["count"] += 1

        # ── permutation combinations (also re-derived so badges are correct)
        combo_map = defaultdict(list)
        for s in all_subs:
            u    = User.query.get(s.user_id)
            f    = _flags(s)
            key  = (
                f["has_padyapuja"],
                f["has_seva_mahaprasad"],
                f["seva_time"] or "none",
                f["has_shejarti"],
            )
            combo_map[key].append(_submission_dict(s, u))

        combinations = []
        for (padya, seva_mp, time, shejarti), users in sorted(combo_map.items(), key=lambda x: -len(x[1])):
            combinations.append({
                "has_padyapuja":       padya,
                "has_seva_mahaprasad": seva_mp,
                "seva_time":           time if time != "none" else None,
                "has_shejarti":        shejarti,
                "count":               len(users),
                "users":               users,
            })

        return jsonify({
            "total":        total,
            "shortlisted":  shortlisted,
            "finalized":    finalized,
            "seva_counts":  dict(seva_counts),
            "time_counts":  dict(time_counts),
            "area_counts":  dict(area_counts),
            "route_counts": sorted(route_counts.values(), key=lambda x: x["route_number"]),
            "combinations": combinations,
        }), 200
    except Exception as e:
        logging.exception("get_summary_admin error: %s", e)
        return jsonify({"error": "Failed to build summary"}), 500


# ─── Admin: shortlist workflow ─────────────────────────────────────────────────

@router.route("/adhik-maas/submissions/<int:submission_id>/shortlist", methods=["PUT"])
def toggle_shortlist(submission_id):
    """
    [Admin] Shortlist or remove from shortlist.
    Body: { "shortlisted": true|false, "admin_notes": "..." (optional) }
    """
    err, status = _require_admin()
    if err is not None:
        return err, status
    from model import db, AdhikMaasSubmission

    data       = request.get_json(force=True, silent=True) or {}
    shortlisted = data.get("shortlisted", True)

    submission = AdhikMaasSubmission.query.get(submission_id)
    if not submission:
        return jsonify({"error": "Submission not found"}), 404

    submission.is_shortlisted  = bool(shortlisted)
    submission.shortlisted_at  = datetime.utcnow() if shortlisted else None
    if "admin_notes" in data:
        submission.admin_notes = data["admin_notes"]

    try:
        db.session.commit()
        return jsonify({
            "message":         "Shortlist updated",
            "id":              submission.id,
            "is_shortlisted":  submission.is_shortlisted,
            "shortlisted_at":  submission.shortlisted_at.isoformat() if submission.shortlisted_at else None,
        }), 200
    except Exception as e:
        db.session.rollback()
        logging.exception("shortlist toggle error: %s", e)
        return jsonify({"error": "Failed to update shortlist"}), 500


@router.route("/adhik-maas/shortlisted", methods=["GET"])
def list_shortlisted():
    """[Admin] Return only shortlisted submissions."""
    err, status = _require_admin()
    if err is not None:
        return err, status
    try:
        from model import AdhikMaasSubmission, User
        subs = AdhikMaasSubmission.query.filter_by(is_shortlisted=True).order_by(AdhikMaasSubmission.area).all()
        out  = [_submission_dict(s, User.query.get(s.user_id)) for s in subs]
        return jsonify({"shortlisted": out, "total": len(out)}), 200
    except Exception as e:
        logging.exception("list_shortlisted error: %s", e)
        return jsonify({"error": "Failed to fetch shortlisted submissions"}), 500


# ─── Admin: finalize workflow ──────────────────────────────────────────────────

@router.route("/adhik-maas/submissions/<int:submission_id>/finalize", methods=["PUT"])
def toggle_finalize(submission_id):
    """
    [Admin] Finalize or un-finalize a user.
    Body: { "finalized": true|false, "admin_notes": "..." (optional) }
    """
    err, status = _require_admin()
    if err is not None:
        return err, status
    from model import db, AdhikMaasSubmission

    data      = request.get_json(force=True, silent=True) or {}
    finalized = data.get("finalized", True)

    submission = AdhikMaasSubmission.query.get(submission_id)
    if not submission:
        return jsonify({"error": "Submission not found"}), 404

    submission.is_finalized = bool(finalized)
    submission.finalized_at = datetime.utcnow() if finalized else None
    if "admin_notes" in data:
        submission.admin_notes = data["admin_notes"]

    try:
        db.session.commit()
        return jsonify({
            "message":       "Finalize status updated",
            "id":            submission.id,
            "is_finalized":  submission.is_finalized,
            "finalized_at":  submission.finalized_at.isoformat() if submission.finalized_at else None,
        }), 200
    except Exception as e:
        db.session.rollback()
        logging.exception("finalize toggle error: %s", e)
        return jsonify({"error": "Failed to update finalize status"}), 500


@router.route("/adhik-maas/finalized", methods=["GET"])
def list_finalized():
    """[Admin] Return only finalized submissions."""
    err, status = _require_admin()
    if err is not None:
        return err, status
    try:
        from model import AdhikMaasSubmission, User
        subs = AdhikMaasSubmission.query.filter_by(is_finalized=True).order_by(AdhikMaasSubmission.area).all()
        out  = [_submission_dict(s, User.query.get(s.user_id)) for s in subs]
        return jsonify({"finalized": out, "total": len(out)}), 200
    except Exception as e:
        logging.exception("list_finalized error: %s", e)
        return jsonify({"error": "Failed to fetch finalized submissions"}), 500


@router.route("/adhik-maas/public-finalized", methods=["GET"])
def public_list_finalized():
    """
    [Public] Return finalized submissions visible to all users.
    Only accessible when the feature-toggle 'adhik_maas_2026_list_finalized' is TRUE.
    Returns a trimmed payload (no admin workflow fields).
    """
    try:
        from model import AdhikMaasSubmission, User, FeatureToggle

        toggle = FeatureToggle.query.filter_by(toggle_name="adhik_maas_2026_list_finalized").first()
        if not toggle or not toggle.toggle_enabled:
            return jsonify({"error": "List not yet published"}), 403

        subs = (
            AdhikMaasSubmission.query
            .filter_by(is_finalized=True)
            .order_by(AdhikMaasSubmission.route_date, AdhikMaasSubmission.area)
            .all()
        )

        def _public_dict(s):
            u = User.query.get(s.user_id)
            return {
                "id":           s.id,
                "user_name":    f"{u.first_name} {u.last_name}".strip() if u else "",
                "flat_no":      getattr(u, "flat_no", None)       if u else None,
                "full_address": getattr(u, "full_address", None)  if u else None,
                "landmark":     getattr(u, "landmark", None)      if u else None,
                "city":         getattr(u, "city", None)          if u else None,
                "state":        getattr(u, "state", None)         if u else None,
                "area":         s.area,
                "route_number": getattr(s, "route_number", None),
                "route_name":   getattr(s, "route_name", None),
                "route_date":   getattr(s, "route_date").isoformat() if getattr(s, "route_date", None) else None,
                "final_seva":   getattr(s, "final_seva", None),
                "seva_label":   getattr(s, "seva_label", None),
                "finalized_at": s.finalized_at.isoformat() if s.finalized_at else None,
            }

        out = [_public_dict(s) for s in subs]
        return jsonify({"finalized": out, "total": len(out)}), 200
    except Exception as e:
        logging.exception("public_list_finalized error: %s", e)
        return jsonify({"error": "Failed to fetch finalized list"}), 500


# ─── Seva options config ──────────────────────────────────────────────────────

_SEVA_TOGGLE_MAP = {
    "padyapuja":       "adhik_maas_seva_padyapuja",
    "seva_mahaprasad": "adhik_maas_seva_mahaprasad",
    "shejarti":        "adhik_maas_seva_shejarti",
}


def _seva_toggle_enabled(name: str) -> bool:
    """Return the enabled status for a seva feature toggle; defaults to True."""
    from model import FeatureToggle
    t = FeatureToggle.query.filter_by(toggle_name=name).first()
    return t.toggle_enabled if t else True


@router.route("/adhik-maas/seva-options", methods=["GET"])
def get_seva_options():
    """
    [Public] Return which seva options the admin has enabled.

    Response:
    {
      "padyapuja":       true,
      "seva_mahaprasad": true,
      "shejarti":        false
    }
    """
    try:
        return jsonify({
            key: _seva_toggle_enabled(toggle_name)
            for key, toggle_name in _SEVA_TOGGLE_MAP.items()
        }), 200
    except Exception as e:
        logging.exception("get_seva_options error: %s", e)
        return jsonify({"error": "Failed to fetch seva options"}), 500


@router.route("/adhik-maas/seva-options", methods=["POST"])
def update_seva_options():
    """
    [Admin] Enable or disable individual seva options.

    Body (any subset):
    {
      "padyapuja":       true | false,
      "seva_mahaprasad": true | false,
      "shejarti":        true | false
    }

    No admin auth check here — mirrors the pattern of /registration-settings
    and /adhik-maas-settings; access is restricted at the app/UI level.
    """
    from model import db, FeatureToggle

    data    = request.get_json(force=True, silent=True) or {}
    updated = {}

    for key, toggle_name in _SEVA_TOGGLE_MAP.items():
        if key in data:
            new_val = bool(data[key])
            toggle  = FeatureToggle.query.filter_by(toggle_name=toggle_name).first()
            if toggle:
                toggle.toggle_enabled = new_val
            else:
                toggle = FeatureToggle(toggle_name=toggle_name, toggle_enabled=new_val)
                db.session.add(toggle)
            updated[key] = new_val

    if not updated:
        return jsonify({"error": "No valid seva option keys provided"}), 400

    try:
        db.session.commit()
        # return the full current state
        return jsonify({
            key: _seva_toggle_enabled(toggle_name)
            for key, toggle_name in _SEVA_TOGGLE_MAP.items()
        }), 200
    except Exception as e:
        db.session.rollback()
        logging.exception("seva-options update error: %s", e)
        return jsonify({"error": "Failed to update seva options"}), 500


# ─── Admin: map data ──────────────────────────────────────────────────────────

def _geocode_address(address):
    if not address or not str(address).strip():
        return None, None
    try:
        from geopy.geocoders import Nominatim
        from geopy.extra.rate_limiter import RateLimiter
        geolocator  = Nominatim(user_agent="upasana-adhik-maas")
        rate_limited = RateLimiter(geolocator.geocode, min_delay_seconds=1.1)
        location    = rate_limited(str(address).strip())
        if location and location.latitude is not None:
            return float(location.latitude), float(location.longitude)
    except Exception as e:
        logging.warning("Geocode failed for %s: %s", (address or "")[:50], e)
    return None, None


@router.route("/adhik-maas/map-data", methods=["GET"])
def get_map_data():
    """[Admin] Submissions with coordinates for map view."""
    err, status = _require_admin()
    if err is not None:
        return err, status

    try:
        from model import db, AdhikMaasSubmission, User

        submissions = AdhikMaasSubmission.query.order_by(AdhikMaasSubmission.submitted_at.desc()).all()
        out = []
        for s in submissions:
            u = User.query.get(s.user_id)
            if not u:
                continue
            name      = f"{u.first_name} {u.last_name}".strip() or f"User #{s.user_id}"
            seva_type = s.seva_label or s.seva_preference or ""
            address   = getattr(u, "full_address", None)
            lat = lon = None
            if getattr(u, "latitude", None) is not None and getattr(u, "longitude", None) is not None:
                try:
                    lat, lon = float(u.latitude), float(u.longitude)
                except (TypeError, ValueError):
                    pass
            if (lat is None or lon is None) and address:
                lat, lon = _geocode_address(address)
                if lat is not None:
                    try:
                        u.latitude  = lat
                        u.longitude = lon
                        db.session.add(u)
                        db.session.commit()
                    except Exception as e:
                        logging.warning("Failed to save lat/lng for user %s: %s", s.user_id, e)
                        db.session.rollback()
            if lat is None or lon is None:
                lat, lon = 18.5204, 73.8567
            out.append({
                "id":               s.id,
                "user_id":          s.user_id,
                "name":             name,
                "seva_type":        seva_type,
                "seva_preference":  s.seva_preference,
                "has_padyapuja":    bool(s.has_padyapuja),
                "has_seva_mahaprasad": bool(s.has_seva_mahaprasad),
                "seva_time":        s.seva_time,
                "has_shejarti":     bool(s.has_shejarti),
                "is_shortlisted":   bool(s.is_shortlisted),
                "is_finalized":     bool(s.is_finalized),
                "latitude":         round(lat, 6),
                "longitude":        round(lon, 6),
            })
        return jsonify(out), 200
    except Exception as e:
        logging.exception("get_map_data error: %s", e)
        return jsonify({"error": "Failed to build map data"}), 500


# ─── Export endpoint ──────────────────────────────────────────────────────────

_TIME_LABELS = {
    "afternoon": "Afternoon",
    "evening":   "Evening",
    "any":       "Any",
}


@router.route("/adhik-maas/export", methods=["GET"])
def export_submissions():
    """
    [Admin] Download all submissions as CSV or XLSX.

    Query params:
      format   – 'csv' (default) or 'xlsx'
      status   – 'all' (default) | 'shortlisted' | 'finalized'
      search   – optional name / mobile / area substring filter
      admin_user_id or admin_mobile – auth (same as all admin endpoints)
    """
    err, status_code = _require_admin()
    if err is not None:
        return err, status_code

    try:
        import io
        import pandas as pd
        from flask import send_file, Response
        from model import AdhikMaasSubmission, User

        fmt           = request.args.get("format", "csv").lower()
        status_filter = request.args.get("status", "all").lower()
        search        = request.args.get("search", "").strip().lower()

        query = AdhikMaasSubmission.query.join(User, AdhikMaasSubmission.user_id == User.id)
        if status_filter == "shortlisted":
            query = query.filter(AdhikMaasSubmission.is_shortlisted == True)
        elif status_filter == "finalized":
            query = query.filter(AdhikMaasSubmission.is_finalized == True)

        submissions = query.order_by(
            AdhikMaasSubmission.route_number,
            AdhikMaasSubmission.area,
        ).all()

        rows = []
        for s in submissions:
            u = User.query.get(s.user_id)
            if not u:
                continue

            name = " ".join(filter(None, [u.first_name, u.middle_name, u.last_name]))

            if search and not any(
                search in str(v).lower()
                for v in [name, u.mobile_number, s.area, s.route_name, s.route_number]
                if v
            ):
                continue

            address = ", ".join(filter(None, [
                u.flat_no, u.full_address, u.landmark,
                u.area, u.city, u.state, u.pincode,
            ]))

            rows.append({
                "Name":               name,
                "Mobile":             u.mobile_number     or "",
                "Zone":               u.zone_code         or "",
                "Reg. Area":          u.area              or "",
                "Address":            address,
                "Route No.":          s.route_number      or "",
                "Route Name":         s.route_name        or "",
                "Submission Area":    s.area              or "",
                "PIN Code":           s.pin_code          or "",
                "Padyapuja":          "Yes" if s.has_padyapuja        else "No",
                "Seva + Mahaprasad":  "Yes" if s.has_seva_mahaprasad  else "No",
                "Time Preference":    _TIME_LABELS.get(s.seva_time, s.seva_time or "—"),
                "Shejarti & Kaakad":  "Yes" if s.has_shejarti         else "No",
                "Seva Label":         s.seva_label        or "",
                "Route Date":         s.route_date.strftime("%d %b %Y") if s.route_date else "",
                "Final Seva":         s.final_seva        or "",
                "Shortlisted":        "Yes" if s.is_shortlisted  else "No",
                "Finalized":          "Yes" if s.is_finalized    else "No",
                "Finalized At":       s.finalized_at.strftime("%d %b %Y %H:%M") if s.finalized_at else "",
                "Admin Notes":        s.admin_notes       or "",
                "Submitted At":       s.submitted_at.strftime("%d %b %Y %H:%M") if s.submitted_at else "",
            })

        if not rows:
            return jsonify({"error": "No data to export for the selected filter"}), 404

        df = pd.DataFrame(rows)
        date_str     = datetime.utcnow().strftime("%Y-%m-%d")
        status_label = status_filter.capitalize()

        if fmt == "xlsx":
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                sheet_name = status_label[:31]
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                ws = writer.sheets[sheet_name]
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill("solid", fgColor="F15700")
                for cell in ws[1]:
                    cell.font      = Font(bold=True, color="FFFFFF")
                    cell.fill      = header_fill
                    cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    width = max((len(str(cell.value or "")) for cell in col), default=8)
                    ws.column_dimensions[col[0].column_letter].width = min(width + 4, 55)
            output.seek(0)
            return send_file(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"Adhik_Maas_{status_label}_{date_str}.xlsx",
            )
        else:
            output = io.StringIO()
            output.write("\ufeff")          # UTF-8 BOM so Excel opens it cleanly
            df.to_csv(output, index=False)
            return Response(
                output.getvalue(),
                mimetype="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition":
                        f'attachment; filename="adhik_maas_{status_filter}_{date_str}.csv"',
                },
            )
    except Exception as e:
        logging.exception("export_submissions error: %s", e)
        return jsonify({"error": "Failed to export submissions"}), 500


# ─── Day Summary (public) ─────────────────────────────────────────────────────

@router.route("/adhik-maas/day-summary", methods=["GET"])
def adhik_maas_day_summary():
    """
    Public endpoint used by the HomeScreen Adhik Maas Daura card.

    Query params:
      date  – YYYY-MM-DD  (required)

    Returns totals + per-seva counts for all *finalized* submissions
    whose route_date matches the given date.
    """
    from model import AdhikMaasSubmission
    from datetime import date as date_type

    date_str = request.args.get("date", "").strip()
    if not date_str:
        return jsonify({"error": "date is required"}), 400

    try:
        date_obj = date_type.fromisoformat(date_str)
    except ValueError:
        return jsonify({"error": "invalid date format, use YYYY-MM-DD"}), 400

    try:
        subs = (
            AdhikMaasSubmission.query
            .filter(
                AdhikMaasSubmission.route_date == date_obj,
                AdhikMaasSubmission.is_finalized == True,
            )
            .all()
        )

        if not subs:
            return jsonify({
                "date":         date_str,
                "route_number": None,
                "route_name":   None,
                "total":        0,
                "sevas": {
                    "padyapuja":          0,
                    "abhishek_afternoon": 0,
                    "abhishek_evening":   0,
                    "shejarti":           0,
                    "mahaprasad":         0,
                },
            }), 200

        route_number = subs[0].route_number
        route_name   = subs[0].route_name

        padyapuja          = sum(1 for s in subs if s.has_padyapuja)
        abhishek_afternoon = sum(1 for s in subs if s.seva_time in ("afternoon",))
        abhishek_evening   = sum(1 for s in subs if s.seva_time in ("evening", "any"))
        shejarti           = sum(1 for s in subs if s.has_shejarti)
        mahaprasad         = sum(1 for s in subs if s.has_seva_mahaprasad)

        return jsonify({
            "date":         date_str,
            "route_number": route_number,
            "route_name":   route_name,
            "total":        len(subs),
            "sevas": {
                "padyapuja":          padyapuja,
                "abhishek_afternoon": abhishek_afternoon,
                "abhishek_evening":   abhishek_evening,
                "shejarti":           shejarti,
                "mahaprasad":         mahaprasad,
            },
        }), 200
    except Exception as e:
        logging.exception("adhik_maas_day_summary error: %s", e)
        return jsonify({"error": "Failed to fetch day summary"}), 500
